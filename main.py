import yaml
import time
import queue
import re
from threading import Thread
from pprint import pformat
from openpyxl import load_workbook, Workbook
from getpass import getpass
from sys import argv
from datetime import datetime
from pathlib import Path
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException


#######################################################################################
# ------------------------------ classes part ----------------------------------------#
#######################################################################################


class CellSiteGateway:
    def __init__(self, ip, host):
        self.hostname = host
        self.ip_address = ip
        self.ssh_conn = None
        self.os_type = "cisco_ios"

        self.connection_status = True  # failed connection status, False if connection fails
        self.connection_error_msg = ""  # connection error message

        self.show_isis_log = ""
        self.show_mac_log = ""
        self.show_arp_log = ""
        self.show_description_log = ""
        self.show_isis_neighbors_log = ""
        self.show_tengig_bw_log = ""
        self.show_tengig_bw = None     # False: 1G, True: 10G

        self.pagg = "-"
        self.exclude_inf = []  # exclude interface vlans
        self.description_exclude = ["UPLINK", "DOWNLINK", "csg", "pagg", "ACCESS", "MGMT", "MNG", "ME"]
        self.bs = {}  # mac : {"bs_id":,"port":, "if_vlan": [], "vlan": []}
        # {48fd.8e05.6fa7: {port: Gi0/8, inf: [1000-1005], bs: AL7410}}
        self.port_bs = {}  # {port: [bs1, bs2]}       {Gi0/8: [AL7374_7008_7007, ALR746]}
        self.ifvlan_bs = {}  # {int vlan: [bs1, bs2]}    {1000: [AL7374_7008_7007, ALR746]}
        self.lag = {}   # Po1 : [members : Gi0/4 Gi0/5, tag: AU7070 Po1, AU7070 Po1]
        self.removed_info = []  # from def delete_info
        
        self.commands = []
        self.configuration_log = []

    def show_commands(self):
        self.show_arp_log = self.ssh_conn.send_command(r"show ip arp vrf MA | exclude -|Incomplete")
        self.show_isis_log = self.ssh_conn.send_command(r"show isis hostname | include pagg")
        self.show_isis_neighbors_log = self.ssh_conn.send_command(r"show isis neighbors | include Vl")
        self.show_mac_log = self.ssh_conn.send_command(r"show mac-address-table")
        self.show_description_log = self.ssh_conn.send_command(r"show interfaces description")
        self.show_tengig_bw_log = self.ssh_conn.send_command(r"show interfaces te0/0 | in BW")

    def parse(self, dev, bs_dict):
        csg_mac_log_parse(dev, bs_dict)
        csg_arp_log_parse(dev)
        csg_define_pagg(dev)
        csg_description_parse(dev)
        csg_tengig_bw_parse(dev)

    def delete_info(self, dev):
        csg_delete_info(dev)
        
    def define_port_bs(self, dev):
        csg_port_bs(dev)
        
    def lag_member_tag(self, dev):
        csg_lag_member_tag(dev)        

    def make_config(self, dev):
        csg_make_config(dev)

    def configure(self, cmd_list):
        self.configuration_log.append(self.ssh_conn.send_config_set(cmd_list))

    def commit(self):
        try:
            self.configuration_log.append(self.ssh_conn.save_config())
        except Exception as err_msg:
            self.configuration_log.append(f"COMMIT is OK after msg:{err_msg}")
            self.configuration_log.append(self.ssh_conn.send_command("\n", expect_string=r"#"))

    def reset(self):
        self.connection_status = True  # failed connection status, False if connection fails
        self.connection_error_msg = ""  # connection error message
        self.show_isis_log = ""
        self.show_mac_log = ""
        self.show_arp_log = ""
        self.show_description_log = ""
        self.show_isis_neighbors_log = ""
        self.show_tengig_bw_log = ""
        self.show_tengig_bw = None
        self.pagg = "-"
        self.exclude_inf = []
        self.bs = {}
        self.port_bs = {}
        self.ifvlan_bs = {}
        self.lag = {}
        self.removed_info = []
        self.commands = []
        self.configuration_log = []


class PaggXR(CellSiteGateway):

    def __init__(self, ip, host):
        CellSiteGateway.__init__(self, ip, host)
        self.os_type = "cisco_xr"

    def commit(self):
        self.configuration_log.append(self.ssh_conn.commit())
        self.ssh_conn.exit_config_mode()

    def configure(self, cmd):
        self.ssh_conn.send_config_set(cmd)
        self.configuration_log.append(self.ssh_conn.send_command("show configuration"))

    def show_commands(self):
        self.show_arp_log = self.ssh_conn.send_command(r"show arp vrf MA | exclude Interface")
        self.show_description_log = self.ssh_conn.send_command(r'show interfaces description')

    def parse(self, dev, bs_dict):
        pagg_arp_log_parse(dev, bs_dict)
        pagg_description_parse(dev)

    def lag_member_tag(self, dev):
        pagg_lag_member_tag(dev)

    def delete_info(self, dev):
        pagg_delete_info(dev)

    def define_port_bs(self, dev):
        pagg_port_bs(dev)

    def make_config(self, dev):
        pagg_make_config(dev)

    def reset(self):
        self.connection_status = True  # failed connection status, False if connection fails
        self.connection_error_msg = ""  # connection error message
        self.show_isis_log = ""
        self.show_mac_log = ""
        self.show_arp_log = ""
        self.show_description_log = ""
        self.show_isis_neighbors_log = ""
        self.exclude_inf = []
        self.bs = {}
        self.port_bs = {}
        self.ifvlan_bs = {}
        self.lag = {}
        self.commands = []
        self.configuration_log = []


class PaggXE(CellSiteGateway):

    def __init__(self, ip, host):
        CellSiteGateway.__init__(self, ip, host)
        self.os_type = "cisco_xe"
        
    def show_commands(self):
        self.show_arp_log = self.ssh_conn.send_command(r"show ip arp vrf MA | exclude -|Incomplete")
        self.show_mac_log = self.ssh_conn.send_command(r"show mac-address-table dynamic")
        self.show_description_log = self.ssh_conn.send_command(r"show interfaces description")

    def parse(self, dev, bs_dict):
        xe_mac_log_parse(dev, bs_dict)
        xe_description_parse(dev)

    def delete_info(self, dev):
        csg_delete_info(dev)
        
    def define_port_bs(self, dev):
        csg_port_bs(dev)
        
    def lag_member_tag(self, dev):
        xe_lag_member_tag(dev)        

    def make_config(self, dev):
        xe_make_config(dev)

    def configure(self, cmd_list):
        self.configuration_log.append(self.ssh_conn.send_config_set(cmd_list))

    def commit(self):
        self.configuration_log.append(self.ssh_conn.save_config())

    def reset(self):
        self.connection_status = True  # failed connection status, False if connection fails
        self.connection_error_msg = ""  # connection error message
        self.show_isis_log = ""
        self.show_mac_log = ""
        self.show_arp_log = ""
        self.show_description_log = ""
        self.show_isis_neighbors_log = ""
        self.show_tengig_bw_log = ""
        self.show_tengig_bw = None
        self.pagg = "-"
        self.exclude_inf = []
        self.bs = {}
        self.port_bs = {}
        self.ifvlan_bs = {}
        self.lag = {}
        self.removed_info = []
        self.commands = []
        self.configuration_log = []   
        
        
        

#######################################################################################
# ------------------------------ def function part -----------------------------------#
#######################################################################################


def get_argv(arguments):
    settings = {"maxth": 10, "conf": False, "os_type": "cisco_ios"}
    mt_pattern = re.compile(r"mt([0-9]+)")
    for arg in arguments:
        if "mt" in arg:
            match = re.search(mt_pattern, arg)
            if match and int(match[1]) <= 100:
                settings["maxth"] = int(match[1])
        elif arg == "cfg" or arg == "CFG" or arg == "conf":
            settings["conf"] = True
        elif arg == "xr" or arg == "XR":
            settings["os_type"] = "cisco_xr"
        elif arg == "xe" or arg == "XE":
            settings["os_type"] = "cisco_xe"    
            
    print()
    print(f"max threads: {settings['maxth']}  configuration mode: {settings['conf']}  ios: {settings['os_type']}\n")
    return settings


def get_user_pw():
    user = input("Enter login: ")
    psw = getpass()
    return user, psw


def get_devinfo(yaml_file, args):
    devs = []
    with open(yaml_file, "r") as file:
        devices_info = yaml.load(file, yaml.SafeLoader)
        if args["os_type"] == "cisco_ios":
            for hostname, ip_address in devices_info.items():
                dev = CellSiteGateway(ip=ip_address, host=hostname)
                devs.append(dev)
        elif args["os_type"] == "cisco_xr":
            for hostname, ip_address in devices_info.items():
                dev = PaggXR(ip=ip_address, host=hostname)
                devs.append(dev)
        elif args["os_type"] == "cisco_xe":
            for hostname, ip_address in devices_info.items():
                dev = PaggXE(ip=ip_address, host=hostname)
                devs.append(dev)
    print()
    return devs


def write_logs(devs, current_time, log_folder, export_device_info, export_excel, settings):
    failed_conn_count = 0
    export_excel(devs, current_time, log_folder)

    conn_msg_filename = log_folder / f"{current_time}_connection_error_msg.txt"
    conn_msg_filename_file = open(conn_msg_filename, "w")
    device_info_filename = log_folder / f"{current_time}_device_info.txt"
    device_info_filename_file = open(device_info_filename, "w")
    config_filename = log_folder / f"{current_time}_configuration_log.txt"
    config_filename_file = open(config_filename, "w")
    commands_filename = log_folder / f"{current_time}_configuration_needed.txt"
    commands_filename_file = open(commands_filename, "w")
    removed_filename = log_folder / f"{current_time}_removed_info.txt"
    removed_filename_file = open(removed_filename, "w")

    for dev in devs:
        if dev.connection_status:
            export_device_info(dev, device_info_filename_file)  # export device info: show, status, etc
        else:
            failed_conn_count += 1
            conn_msg_filename_file.write("-" * 80 + "\n")
            conn_msg_filename_file.write(f"### {dev.hostname} : {dev.ip_address} ###\n\n")
            conn_msg_filename_file.write(f"{dev.connection_error_msg}\n")
            config_filename_file.write("\n\n")
            
        if settings["conf"] and dev.commands:
            config_filename_file.write("#" * 80 + "\n")
            config_filename_file.write(f"### {dev.hostname} : {dev.ip_address} ###\n\n")
            config_filename_file.write("".join(dev.configuration_log))
            config_filename_file.write("\n\n")
        elif not settings["conf"] and dev.commands:
            commands_filename_file.write(f"### {dev.hostname} : {dev.ip_address}\n\n")
            commands_filename_file.write("\n".join(dev.commands))
            commands_filename_file.write("\n\n\n")
            
        if dev.removed_info:
            removed_filename_file.write(f"{dev.hostname}\t{' '.join(dev.removed_info)}\n")

    conn_msg_filename_file.close()
    device_info_filename_file.close()
    config_filename_file.close()
    commands_filename_file.close()
    removed_filename_file.close()

    if not settings["conf"]:
        config_filename.unlink()
    if all([dev.connection_status is True for dev in devs]):
        conn_msg_filename.unlink()

    return failed_conn_count


#######################################################################################
# ------------------------------ get bs port -----------------------------------------#
#######################################################################################


def load_excel(curr_date, curr_time):
    excel_file = input("Enter IP-MAC excel file (by default no excel file is loaded): ")
    print()
    result = {}  # mac : bs
    if len(excel_file) > 0:
        wb = load_workbook(excel_file)
        first_sheet = wb.sheetnames[0]
        sheet = wb[first_sheet]
        x = 2

        while True:
            mac = sheet.cell(row=x, column=4).value
            bs = sheet.cell(row=x, column=1).value
            if bs:
                x += 1
                mac_split = mac.split(":")
                mac_final = "{}{}.{}{}.{}{}".format(mac_split[0], mac_split[1],
                                                    mac_split[2], mac_split[3],
                                                    mac_split[4], mac_split[5])
                result[mac_final] = bs
            else:
                break

        with open("mac_bs.yaml", "w") as output_file:
            output_file.write(f"# {curr_date} {curr_time}\n\n")
            for i, j in result.items():
                output_file.write(f"{i} : {j}\n")

    else:
        with open("mac_bs.yaml", "r") as file:
            yaml_file = yaml.load(file, yaml.SafeLoader)
            result.update(yaml_file)

    return result
    

def export_excel(devs, current_time, log_folder):
    filename = log_folder / f"{current_time}_mbh_bs_list.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.append(["PAGG",
                  "CSG hostname",
                  "CSG loopback0",
                  "CSG port",
                  "port tag",
                  "BS",
                  "comments"])
    for dev in devs:
        if dev.connection_status:
            for port, port_info in dev.port_bs.items():
                if dev.lag.get(port):
                    sheet.append([dev.pagg,
                                  dev.hostname,
                                  dev.ip_address,
                                  f'{port} ({len(dev.lag[port]["members"])} Gps)',
                                  f'{" ".join(set(dev.lag[port]["tag"]))}',
                                  ' '.join(port_info["bs"])])
                else:
                    if "Te" in port and dev.show_tengig_bw == "1G":
                        sheet.append([dev.pagg,
                                      dev.hostname,
                                      dev.ip_address,
                                      f'{port} (1 Gps)', 
                                      port_info["tag"],
                                      ' '.join(port_info["bs"])])
                    else:
                        sheet.append([dev.pagg,
                                      dev.hostname,
                                      dev.ip_address,
                                      port, 
                                      port_info["tag"],
                                      ' '.join(port_info["bs"])])
        else:
            sheet.append([dev.pagg,
                          dev.hostname,
                          dev.ip_address,
                          "-",
                          "-",
                          "-",
                          "unavailable"])

    wb.save(filename)


def export_device_info(dev, export_file):
    export_file.write("#" * 80 + "\n")
    export_file.write(f"### {dev.hostname} : {dev.ip_address} ###\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.show_isis_log\n\n")
    export_file.write(dev.show_isis_log)
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.show_mac_log\n\n")
    export_file.write(dev.show_mac_log)
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.show_arp_log\n\n")
    export_file.write(dev.show_arp_log)
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.show_description_log\n\n")
    export_file.write(dev.show_description_log)
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.show_isis_neighbors_log\n\n")
    export_file.write(dev.show_isis_neighbors_log)
    export_file.write("\n\n")
    
    export_file.write("-" * 80 + "\n")
    export_file.write("device.bs\n\n")
    export_file.write(pformat(dev.bs))
    export_file.write("\n\n")
    
    export_file.write("-" * 80 + "\n")
    export_file.write("device.port_bs\n\n")
    export_file.write(pformat(dev.port_bs))
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.ifvlan_bs\n\n")
    export_file.write(pformat(dev.ifvlan_bs))
    export_file.write("\n\n")

    export_file.write("-" * 80 + "\n")
    export_file.write("device.lag\n\n")
    export_file.write(pformat(dev.lag))
    export_file.write("\n\n")
    
    export_file.write("-" * 80 + "\n")
    export_file.write("device.exclude_inf\n\n")
    export_file.write(pformat(dev.exclude_inf))
    export_file.write("\n\n")
    
    export_file.write("-" * 80 + "\n")
    export_file.write("device.commands\n\n")
    export_file.write(pformat(dev.commands))
    export_file.write("\n\n")


def define_inf_exclude(dev):
    dev.exclude_inf.extend([str(i) for i in range(1080, 1199)])  # MA BGP
    dev.exclude_inf.extend([str(i) for i in range(4000, 4099)])  # MW MGMT
    dev.exclude_inf.extend([str(i) for i in range(2020, 2099)])  # SMART METERING

    for line in dev.show_isis_neighbors_log.splitlines():
        match = re.search(r".*L2 +Vl(\d+) +", line) # akta-040001-pag L2 Vl(200) 10.238.121.65
        if match:
            dev.exclude_inf.append(match[1])


def csg_mac_log_parse(dev, bs_dict):
    pattern = re.compile(r"(\d+)\s+(\w{4}\.\w{4}\.\w{4})\s+DYNAMIC\s+(\S+)")
    # (3001)    (48fd.8e05.6fa7)    DYNAMIC     (Gi0/8)
    for line in dev.show_mac_log.splitlines():
        match = re.search(pattern, line)
        if match:
            vlan = match[1]  # 3001
            mac = match[2]  # 48fd.8e05.6fa7
            port = match[3]  # Gi0/8
            if bs_dict.get(mac):
                bs = bs_dict[mac]
            else:
                bs = mac
                
            if dev.bs.get(mac):
                dev.bs[mac]["vlan"].append(vlan)
            else:
                dev.bs[mac] = {"bs_id": bs,
                               "port": port,
                               "if_vlan": [],
                               "vlan": [vlan]}
                               
            if "Po" in port:
                dev.lag[port] = {"members": [], "tag": []}


def xe_mac_log_parse(dev, bs_dict):
    pattern = re.compile(r"(\d+) +(\w{4}\.\w{4}\.\w{4}) +DYNAMIC +(Gi\d/\d/\d)")
    # (1000)  (18de.d7aa.7264)  DYNAMIC  (Gi0/5/6).Efp1000
    for line in dev.show_mac_log.splitlines():
        match = re.search(pattern, line)
        if match:
            vlan = match[1] # продублировать в if_vlan
            mac = match[2]
            port = match[3]
            
            if bs_dict.get(mac):
                bs = bs_dict[mac]
            else:
                bs = mac
                
            if dev.bs.get(mac):
                dev.bs[mac]["vlan"].append(vlan)
                dev.bs[mac]["if_vlan"].append(vlan)
            else:
                dev.bs[mac] = {"bs_id": bs,
                               "port": port,
                               "if_vlan": [vlan],
                               "vlan": [vlan]}
                               
            if "Po" in port:
                dev.lag[port] = {"members": [], "tag": []}

def csg_arp_log_parse(dev):
    pattern = re.compile(r"Internet\s+\d+\.\d+\.\d+\.\d+\s+\d+\s+(\w{4}\.\w{4}\.\w{4})\s+ARPA\s+Vlan(\d+)")
    # Internet  10.165.161.87          11   (d849.0b95.af44)  ARPA   Vlan(1000)
    for line in dev.show_arp_log.splitlines():
        match = re.search(pattern, line)  # ip mac inf_vlan
        if match:
            mac = match[1]  # d849.0b95.af44
            inf = match[2]  # 1000 (without Vlan)
            if dev.bs.get(mac):
                dev.bs[mac]["if_vlan"].append(inf)
            else:
                print(f"{dev.hostname:39}: arp_log_parse - {mac} not in MAC table")


def pagg_arp_log_parse(dev, bs_dict):
    pattern = re.compile(r"\d+\.\d+\.\d+\.\d+\s+[0-9:]{8}\s+(\w{4}\.\w{4}\.\w{4})\s+Dynamic\s+ARPA\s+"
                         r"([-A-Za-z]+)([0-9/]+)\.(\d+)$")
    # 10.146.56.1     00:02:06   (883f.d304.e2a1)  Dynamic    ARPA  (GigabitEthernet)(0/0/0/5).(1080)
    # 10.164.24.243   00:02:11   (845b.1260.9241)  Dynamic    ARPA  (Bundle-Ether)(10).(1004)

    for line in dev.show_arp_log.splitlines():
        match = re.search(pattern, line)  # ip mac inf_vlan
        if match:
            mac = match[1]              # 883f.d304.e2a1
            port_ethernet = match[2]    # GigabitEthernet
            port_number = match[3]      # 0/0/0/5
            vlan = match[4]             # 1080
            
            if bs_dict.get(mac):
                bs = bs_dict[mac]
            else:
                bs = mac

            if port_ethernet == "Bundle-Ether":
                port_ethernet = "BE"
            elif port_ethernet == "TenGigE":
                port_ethernet = "Te"
            elif port_ethernet == "GigabitEthernet":
                port_ethernet = "Gi"
            else:
                print(f"{dev.hostname:39}:pagg_arp_log_parse: Gi,Te,Be not in {port_ethernet}")

            if dev.bs.get(mac):
                dev.bs[mac]["if_vlan"].append(vlan)
                dev.bs[mac]["vlan"].append(vlan)
            else:
                dev.bs[mac] = {"port": f'{port_ethernet}{port_number}',
                               "if_vlan": [vlan],
                               "vlan": [vlan],
                               "bs_id": bs}
                               
            if port_ethernet == "BE":
                dev.lag[f"{port_ethernet}{port_number}"] = {"members": [], "tag": []}
    

def csg_description_parse(dev):
    pattern_port = re.compile(r"((?:Gi|Te|Po)\S+)\s+up\s+up\s*(.*)")  # (Gi0/6) up up (AK7137 BS: ALG005 AK7160)
    pattern_port_tag_bs = re.compile(r"(?:(.*)\s)?BS:\s?(.*)")  # (AK7137) BS: (ALG005 AK7160)
    pattern_inf = re.compile(r"Vl(\d+)\s+up\s+up\s*(.*)")  # Vl(1000) up up (ABIS BS: ALG005 AK7160)
    pattern_inf_tag_bs = re.compile(r"(?:.*\s)?BS:\s?(.*)")  # ABIS BS: (ALG005 AK7160)
    for line in dev.show_description_log.splitlines():
        match_port = re.search(pattern_port, line)
        match_inf = re.search(pattern_inf, line)
        if match_port:
            port = match_port[1]
            description = match_port[2]
            if not any(i in line for i in dev.description_exclude):
                if "BS:" in description:
                    match_port_tag_bs = re.search(pattern_port_tag_bs, description)
                    if match_port_tag_bs:
                        tag = match_port_tag_bs[1]
                        bs = match_port_tag_bs[2]
                        dev.port_bs[port] = {"tag": f'{tag if tag else ""}',
                                             "bs": [],
                                             "short_bs_descr": "",
                                             "short_bs_from_device": bs,
                                             "bs_from_device": []}
                    else:
                        print(f"{dev.hostname:39}{match_port_tag_bs} re match error")
                else:
                    if len(description) > 0:
                        dev.port_bs[port] = {"tag": description,
                                             "bs": [],
                                             "short_bs_descr": "",
                                             "short_bs_from_device": "",
                                             "bs_from_device": []}
                    else:
                        dev.port_bs[port] = {"tag": "",
                                             "bs": [],
                                             "short_bs_descr": "",
                                             "short_bs_from_device": "",
                                             "bs_from_device": []}
        elif match_inf:
            inf = match_inf[1]
            description = match_inf[2]
            tag = ""
            if inf not in dev.exclude_inf and not any(i in line for i in dev.description_exclude):
                for m in ["ABIS", "IUB", "OAM", "S1U", "S1MME", "X2", "S1C"]:
                    if m in description:
                        tag = m
                        break
                if tag == "":
                    print(f"{dev.hostname:39}no ABIS,X2,IUB,S1MME,S1U,S1C,OAM in description interface vlan{inf}")
                if "BS:" in description:
                    match_inf_tag_bs = re.search(pattern_inf_tag_bs, description)
                    if match_inf_tag_bs:
                        bs = match_inf_tag_bs[1]
                        dev.ifvlan_bs[inf] = {"tag": tag,
                                              "bs": [],
                                              "short_bs_descr": "",
                                              "short_bs_from_device": bs,
                                              "bs_from_device": []}
                    else:
                        print(f"{dev.hostname:39}{match_inf_tag_bs} re match error")
                else:
                    if len(description) > 0 and tag != "":
                        dev.ifvlan_bs[inf] = {"tag": tag,
                                              "bs": [],
                                              "short_bs_descr": "",
                                              "short_bs_from_device": "",
                                              "bs_from_device": []}
                    else:
                        print(f"{dev.hostname:39}no description interface vlan{inf}")


def xe_description_parse(dev):
    pattern_port = re.compile(r"((?:Gi|Po)\S+)\s+up\s+up\s*(.*)")
    pattern_port_tag_bs = re.compile(r"(?:(.*)\s)?BS:\s?(.*)")
    pattern_inf = re.compile(r"BD(\d+)\s+up\s+up\s*(.*)")       # BD(1000)
    pattern_inf_tag_bs = re.compile(r"(?:.*\s)?BS:\s?(.*)")     # ABIS BS: (ALG005 AK7160)
    
    for line in dev.show_description_log.splitlines():
        match_port = re.search(pattern_port, line)
        match_inf = re.search(pattern_inf, line)
        if match_port:
            port = match_port[1]
            description = match_port[2]
            if not any(i in line for i in dev.description_exclude):
                if "BS:" in description:
                    match_port_tag_bs = re.search(pattern_port_tag_bs, description)
                    if match_port_tag_bs:
                        tag = match_port_tag_bs[1]
                        bs = match_port_tag_bs[2]
                        dev.port_bs[port] = {"tag": f'{tag if tag else ""}',
                                             "bs": [],
                                             "short_bs_descr": "",
                                             "short_bs_from_device": bs,
                                             "bs_from_device": []}
                    else:
                        print(f"{dev.hostname:39}{match_port_tag_bs} re match error")
                else:
                    if len(description) > 0:
                        dev.port_bs[port] = {"tag": description,
                                             "bs": [],
                                             "short_bs_descr": "",
                                             "short_bs_from_device": "",
                                             "bs_from_device": []}
                    else:
                        dev.port_bs[port] = {"tag": "",
                                             "bs": [],
                                             "short_bs_descr": "",
                                             "short_bs_from_device": "",
                                             "bs_from_device": []}
        elif match_inf:
            inf = match_inf[1]
            description = match_inf[2]
            tag = ""
            if inf not in dev.exclude_inf and not any(i in line for i in dev.description_exclude):
                for m in ["ABIS", "IUB", "OAM", "S1U", "S1MME", "X2", "S1C"]:
                    if m in description:
                        tag = m
                        break
                if tag == "":
                    print(f"{dev.hostname:39}no ABIS,X2,IUB,S1MME,S1U,S1C,OAM in description interface vlan{inf}")
                if "BS:" in description:
                    match_inf_tag_bs = re.search(pattern_inf_tag_bs, description)
                    if match_inf_tag_bs:
                        bs = match_inf_tag_bs[1]
                        dev.ifvlan_bs[inf] = {"tag": tag,
                                              "bs": [],
                                              "short_bs_descr": "",
                                              "short_bs_from_device": bs,
                                              "bs_from_device": []}
                    else:
                        print(f"{dev.hostname:39}{match_inf_tag_bs} re match error")
                else:
                    if len(description) > 0 and tag != "":
                        dev.ifvlan_bs[inf] = {"tag": tag,
                                              "bs": [],
                                              "short_bs_descr": "",
                                              "short_bs_from_device": "",
                                              "bs_from_device": []}
                    else:
                        print(f"{dev.hostname:39}no description interface vlan{inf}")


def pagg_description_parse(dev):
    pattern = re.compile(r"((?:Gi|Te|BE)[0-9/]+)\s+up\s+up\s*(.*)$")
    # (Gi0/0/0/5)     up  up  (AU7104 BS: ZHA012)
    # Gi0/0/0/5.1000  up  up  AU7104
    pattern_tag_bs = re.compile(r"(?:(.*)\s)?BS:\s?(.*)")
    # (AU7104) BS: (ZHA012)

    for line in dev.show_description_log.splitlines():
        match = re.search(pattern, line)
        if match:
            port = match[1]
            description = match[2]
            if not any(i in line for i in dev.description_exclude):
                if "BS:" in description:
                    match_port_tag_bs = re.search(pattern_tag_bs, description)
                    if match_port_tag_bs:
                        tag = match_port_tag_bs[1]
                        bs = match_port_tag_bs[2]
                        dev.port_bs[port] = {"tag": f'{tag if tag else ""}',
                                             "bs": [],
                                             "short_bs_descr": "",
                                             "short_bs_from_device": bs,
                                             "bs_from_device": []}
                    else:
                        print(f"{dev.hostname:39}{match_port_tag_bs} re match error")
                else:
                    if len(description) > 0:
                        dev.port_bs[port] = {"tag": description,
                                             "bs": [],
                                             "short_bs_descr": "",
                                             "short_bs_from_device": "",
                                             "bs_from_device": []}
                    else:
                        dev.port_bs[port] = {"tag": "",
                                             "bs": [],
                                             "short_bs_descr": "",
                                             "short_bs_from_device": "",
                                             "bs_from_device": []}


def description_bs_parse(dev):
    for port, port_info in dev.port_bs.items():
        if port_info["short_bs_from_device"]:
            for i in port_info["short_bs_from_device"].split():
                if "_" in i:  # AU7311_7069
                    city = i[:2]  # AU
                    for j in i.split("_"):  # [AU7311, 7069]
                        if city in j:
                            dev.port_bs[port]["bs_from_device"].append(j)
                        else:
                            dev.port_bs[port]["bs_from_device"].append(f"{city}{j}")
                else:
                    dev.port_bs[port]["bs_from_device"].append(i)  # ALR100
    for inf, inf_info in dev.ifvlan_bs.items():
        if inf_info["short_bs_from_device"]:
            for i in inf_info["short_bs_from_device"].split():
                if "_" in i:
                    city = i[:2]
                    for j in i.split("_"):
                        if city in j:
                            dev.ifvlan_bs[inf]["bs_from_device"].append(j)
                        else:
                            dev.ifvlan_bs[inf]["bs_from_device"].append(f"{city}{j}")
                else:
                    dev.ifvlan_bs[inf]["bs_from_device"].append(i)


def csg_tengig_bw_parse(dev):
    if "BW 1000000 Kbit" in dev.show_tengig_bw_log:
        dev.show_tengig_bw = "1G"


def csg_lag_member_tag(dev):
    if dev.lag:
        pattern = re.compile(r"((?:Gi|Te)\S+)\s+up\s+up\s*(.*)")
        for port in dev.lag:
            log = dev.ssh_conn.send_command(f"show etherchannel {port[2:]} summary | include LACP") # только номер порта
            lag_members = re.findall(r"(?:Gi|Te)\d/\d{1,2}", log)
            dev.lag[port]["members"].extend(lag_members)
            for line in dev.show_description_log.splitlines():
                match = re.search(pattern, line)
                if match:
                    port_gi = match[1]
                    tag = match[2]
                    if port_gi in lag_members:
                       dev.lag[port]["tag"].append(tag)
    
    
def xe_lag_member_tag(dev):
    if dev.lag:
        pattern = re.compile(r"((?:Gi)\S+)\s+up\s+up\s*(.*)")
        for port in dev.lag:
            log = dev.ssh_conn.send_command(f"show etherchannel summary | include {port}")
            lag_members = re.findall(r"(?:Gi)\d/\d/\d", log)
            dev.lag[port]["members"].extend(lag_members)
            for line in dev.show_description_log.splitlines():
                match = re.search(pattern, line)
                if match:
                    port_gi = match[1]
                    tag = match[2]
                    if port_gi in lag_members:
                       dev.lag[port]["tag"].append(tag)
            

def pagg_lag_member_tag(dev):
    if dev.lag:
        pattern = re.compile(r"(Gi[0-9/]+)\s+up\s+up\s*(.*)$")
        for port in dev.lag:
            log = dev.ssh_conn.send_command(f"show bundle {port} | include Local") 
            lag_members = re.findall(r"(?:Gi|Te)\d/\d/\d/\d{1,2}", log)
            dev.lag[port]["members"].extend(lag_members)
            for line in dev.show_description_log.splitlines():
                match = re.search(pattern, line)
                if match:
                    port_gi = match[1]
                    tag = match[2]
                    if port_gi in lag_members:
                       dev.lag[port]["tag"].append(tag)
                        

def csg_delete_info(dev):
    delete_mac = []
    dev_exclude = ["shor.asta-032001-csg-1", "esil.koks-025001-csg-1"]  # устройство на которых БС прописана как L2 
    
    for mac, bs_info in dev.bs.items():
        if len(bs_info["bs_id"]) == 14:     # удалить все неопределенные MAC, 0046.4bb4.8f76=14
            delete_mac.append(mac)
            if mac in dev.show_arp_log and len(bs_info["vlan"]) > 2:
                print(f"{dev.hostname:39}csg_delete_info: {mac} not in MAC-BS.excel table")
        else:
            if mac not in dev.show_arp_log and dev.hostname not in dev_exclude:
                delete_mac.append(mac)

    if delete_mac:
        for i in set(delete_mac):
            dev.removed_info.extend(dev.bs[i]['vlan'])
            del dev.bs[i]
            
    if dev.lag:
        for lag_info in dev.lag.values():
            for port in lag_info["members"]:
                if dev.port_bs.get(port):
                    del dev.port_bs[port]


def pagg_delete_info(dev):
    delete_mac = []
    
    for mac, bs_info in dev.bs.items():
        if len(bs_info["bs_id"]) == 14:     # удалить все неопределенные MAC, 0046.4bb4.8f76=14
            delete_mac.append(mac)
            if len(bs_info["vlan"]) > 2:
                print(f"{dev.hostname:39}csg_delete_info: {mac} not in MAC-BS.excel table")

    if delete_mac:
        for i in set(delete_mac):
            dev.removed_info.extend(dev.bs[i]['vlan'])
            del dev.bs[i]
            
    if dev.lag:
        for lag_info in dev.lag.values():
            for port in lag_info["members"]:
                if dev.port_bs.get(port):
                    del dev.port_bs[port]


def csg_define_pagg(dev):
    pattern = re.compile(r"[0-9.]{14} ([a-z.]+-\d+-pagg-\d)")
    for line in dev.show_isis_log.splitlines():
        match = re.search(pattern, line)
        if match:
            dev.pagg = match[1]


def csg_port_bs(dev):
    for bs_info in dev.bs.values():
        port = bs_info["port"]
        bs = bs_info["bs_id"]
        ifvlanlist = bs_info["if_vlan"]
        if dev.port_bs.get(port):
            dev.port_bs[port]["bs"].append(bs)
        else:
            print(f"{dev.hostname:39}{port} not in port_bs dict")
        for ifvlan in ifvlanlist:
            if dev.ifvlan_bs.get(ifvlan):
                dev.ifvlan_bs[ifvlan]["bs"].append(bs)
            else:
                print(f"{dev.hostname:39}{ifvlan} vlan not in ifvlan_bs dict")


def pagg_port_bs(dev):
    for bs_info in dev.bs.values():
        port = bs_info["port"]
        bs = bs_info["bs_id"]
        if dev.port_bs.get(port):
            dev.port_bs[port]["bs"].append(bs)
        else:
            print(f"{dev.hostname:39} port not in port_bs dict")


def shorten_bs(dev):
    pattern = re.compile(r"^([A-Z]{2})(\d{4})")  # (AL)(7374)
    for port, port_info in dev.port_bs.items():
        city_bs = {"others": []}  # AL:[7341,7000] AS:[7007,7000] others:[ALR734,TEST_BS]
        bs_desc = []  # [AL7374_7008_7007, AS7374_7375, ALR746]       
        for bs in port_info["bs"]:
            match = re.search(pattern, bs)
            if match:
                region = match[1]  # AL
                bs_number = match[2]  # 7341
                if city_bs.get(region):
                    city_bs[region].append(bs_number)  # AL: [7374, 7000], AS: [7007, 7008]
                else:
                    city_bs[region] = [bs_number]  # AL: [7374], AS: [7007]
            else:
                city_bs["others"].append(bs)  # "others": [ALR734, AL100, TEST_BS]
        for i, j in city_bs.items():
            if i != "others":
                bs_desc.append(f"{i}{'_'.join(j)}")  # ["AL7374_7008_7007"]
            elif i == "others" and len(j) > 0:
                bs_desc.extend(j)  # ["AL7374_7008_7007", "ALR734", "AL100", "TEST_BS"]
        dev.port_bs[port]["short_bs_descr"] = " ".join(bs_desc)

    for inf, inf_info in dev.ifvlan_bs.items():
        city_bs = {"others": []}
        bs_desc = []
        for bs in inf_info["bs"]:
            match = re.search(pattern, bs)
            if match:
                region = match[1]
                bs_number = match[2]
                if city_bs.get(region):
                    city_bs[region].append(bs_number)
                else:
                    city_bs[region] = [bs_number]
            else:
                city_bs["others"].append(bs)
        for i, j in city_bs.items():
            if i != "others":
                bs_desc.append(f"{i}{'_'.join(j)}")
            elif i == "others" and len(j) > 0:
                bs_desc.extend(j)

        dev.ifvlan_bs[inf]["short_bs_descr"] = " ".join(bs_desc)


def csg_make_config(dev):
    for port, port_info in dev.port_bs.items():
        if set(port_info["bs"]) != set(port_info["bs_from_device"]):
            dev.commands.append(f"interface {port}")
            dev.commands.append(f"description {port_info['tag']} BS: {port_info['short_bs_descr']}")

    for inf, inf_info in dev.ifvlan_bs.items():
        if set(inf_info["bs"]) != set(inf_info["bs_from_device"]):
            dev.commands.append(f"interface Vlan{inf}")
            dev.commands.append(f"description {inf_info['tag']} BS: {inf_info['short_bs_descr']}")


def xe_make_config(dev):
    for port, port_info in dev.port_bs.items():
        if set(port_info["bs"]) != set(port_info["bs_from_device"]):
            dev.commands.append(f"interface {port}")
            dev.commands.append(f"description {port_info['tag']} BS: {port_info['short_bs_descr']}")

    for inf, inf_info in dev.ifvlan_bs.items():
        if set(inf_info["bs"]) != set(inf_info["bs_from_device"]):
            dev.commands.append(f"interface BDI{inf}")
            dev.commands.append(f"description {inf_info['tag']} BS: {inf_info['short_bs_descr']}")


def pagg_make_config(dev):
    for port, port_info in dev.port_bs.items():
        if set(port_info["bs"]) != set(port_info["bs_from_device"]):
            dev.commands.append(f"interface {port} description {port_info['tag']} BS: {port_info['short_bs_descr']}")


def too_long_description(dev):
    for cmd in dev.commands:
        if len(cmd) > 211:
            print(f"{dev.hostname:39}description is longer than 200")


def configure(dev, settings):
    if settings["conf"]:
        if len(dev.commands) > 0:
            dev.configure(dev.commands)
            dev.commit()
        else:
            print(f"{dev.hostname:39}cfg is not needed")
    else:
        if len(dev.commands) > 0:
            print(f"{dev.hostname:39}cfg is needed")


#######################################################################################
# ------------------------------              ----------------------------------------#
#######################################################################################

def connect_dev(my_username, my_password, dev_queue, bs_dict, settings):
    while True:
        dev = dev_queue.get()
        i = 0
        while True:
            try:
                # print(f"{device.hostname:23}{device.ip_address:16}")
                dev.ssh_conn = ConnectHandler(device_type=dev.os_type, ip=dev.ip_address,
                                              username=my_username, password=my_password)
                dev.show_commands()
                define_inf_exclude(dev)
                dev.parse(dev, bs_dict)
                dev.lag_member_tag(dev)
                dev.delete_info(dev)
                description_bs_parse(dev)
                dev.define_port_bs(dev)
                shorten_bs(dev)
                dev.make_config(dev)
                too_long_description(dev)
                configure(dev, settings)
                dev.ssh_conn.disconnect()
                dev_queue.task_done()
                break

            except NetMikoTimeoutException as err_msg:
                dev.connection_status = False
                dev.connection_error_msg = str(err_msg)
                print(f"{dev.hostname:23}{dev.ip_address:16}timeout")
                dev_queue.task_done()
                break

            except Exception as err_msg:
                if i == 1:  # tries
                    dev.connection_status = False
                    dev.connection_error_msg = str(err_msg)
                    print(f"{dev.hostname:23}{dev.ip_address:16}{'BREAK connection failed':20} i={i}")
                    dev_queue.task_done()
                    break
                else:
                    i += 1
                    dev.reset()
                    print(f"{dev.hostname:23}{dev.ip_address:16}{'ERROR connection failed':20} i={i}")
                    time.sleep(5)


#######################################################################################
# ------------------------------ test        -----------------------------------------#
#######################################################################################

def test_connect_dev(dev, settings):
    if settings["os_type"] == "cisco_ios":
        with open("test_arp.txt", "r") as arp:
            dev.show_arp_log = arp.read()
        with open("test_descrip.txt", "r") as descr:
            dev.show_description_log = descr.read()
        with open("test_isis_pagg.txt", "r") as isis_host:
            dev.show_isis_log = isis_host.read()
        with open("test_isis_neig.txt", "r") as isis_neigh:
            dev.show_isis_neighbors_log = isis_neigh.read()
        with open("test_mac.txt", "r") as mac:
            dev.show_mac_log = mac.read()

    elif settings["os_type"] == "cisco_xr":
        with open("test_pagg_arp.txt", "r") as arp:
            dev.show_arp_log = arp.read()
        with open("test_pagg_description.txt", "r") as descr:
            dev.show_description_log = descr.read()


def test_connect(dev_queue, settings):
    dev = dev_queue.get()
    test_connect_dev(dev, settings)
    dev.show_commands()
    define_inf_exclude(dev)
    dev.parse(dev, bs_dict)
    dev.lag_member_tag(dev)
    dev.delete_info(dev)
    description_bs_parse(dev)
    dev.define_port_bs(dev)
    shorten_bs(dev)
    dev.make_config(dev)
    too_long_description(dev)
    dev_queue.task_done()
    
def test_connect2(my_username, my_password, dev_queue, bs_dict, settings):
    dev = dev_queue.get()
    dev.ssh_conn = ConnectHandler(device_type=dev.os_type, ip=dev.ip_address, username=my_username, password=my_password)
    dev.show_commands()
    define_inf_exclude(dev)
    dev.parse(dev, bs_dict)
    dev.lag_member_tag(dev)
    dev.delete_info(dev)
    description_bs_parse(dev)
    dev.define_port_bs(dev)
    shorten_bs(dev)
    dev.make_config(dev)
    too_long_description(dev)
    dev_queue.task_done()


#######################################################################################
# ------------------------------ main part -------------------------------------------#
#######################################################################################

starttime = datetime.now()
current_date = starttime.strftime("%Y.%m.%d")
current_time = starttime.strftime("%H.%M.%S")

log_folder = Path(f"{Path.cwd()}/logs/{current_date}/")  # current dir / logs / date /
log_folder.mkdir(exist_ok=True)

q = queue.Queue()

argv_dict = get_argv(argv)
username, password = get_user_pw()
devices = get_devinfo("devices.yaml", argv_dict)
mac_bs = load_excel(current_date, current_time)  # 04bd.70dc.a7ee : TA7175, информация от МТС

total_devices = len(devices)

print("-------------------------------------------------------------------------------------------------------")
print("hostname               ip address      comment")
print("---------------------- --------------- ----------------------------------------------------------------")

for i in range(argv_dict["maxth"]):
    thread = Thread(target=connect_dev, args=(username, password, q, mac_bs, argv_dict))
    # thread = Thread(target=test_connect, args=(q, mac_bs, argv_dict))
    # thread = Thread(target=test_connect2, args=(username, password, q, mac_bs, argv_dict))
    thread.setDaemon(True)
    thread.start()

for device in devices:
    q.put(device)

q.join()

print()
failed_connection_count = write_logs(devices, current_time, log_folder, export_device_info, export_excel, argv_dict)
duration = datetime.now() - starttime

print("-------------------------------------------------------------------------------------------------------")
print(f"failed connection: {failed_connection_count}  total device number: {total_devices}")
print(f"elapsed time: {duration}")
print("-------------------------------------------------------------------------------------------------------")
